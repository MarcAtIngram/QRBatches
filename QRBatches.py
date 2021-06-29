import re
from openpyxl import load_workbook
import pyqrcode
import png

wb = load_workbook(filename='trailers.xlsx')
ws = wb["Sheet1"]

for row in ws.iter_rows(min_row=1, max_col=5, values_only=True):
    # if this code feels verbose, it is! :D
    # most steps rely on both the Excel sheet and Youtube links to remain the same, but they may not
    # we're verbose here in the hope that the code is easier to maintain

    movieName = re.sub('[^A-Za-z0-9 ]+', '', row[0])
    youtubeLongLink = str(row[4]).strip()

    assetStart = youtubeLongLink.find("=")+1
    linkLength = len(youtubeLongLink)
    assetID = youtubeLongLink[assetStart:linkLength].strip()
    youtubeShortLink = "https://youtu.be/"+assetID
    qrFileName = "Trailer QR - "+movieName+" - "+assetID+".png"

    qrEncoding = pyqrcode.create(youtubeShortLink, error='L')
    qrEncoding.png(qrFileName, scale=5)
